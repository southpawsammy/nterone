import pandas as pd 
import numpy as np
import pathlib
import numbers
from openpyxl import load_workbook
from openpyxl.styles import DEFAULT_FONT
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

#takes a dataframe and reformats to show subtotal breakdown
def reformat(df):
    project_name = df['Project'].iloc[0]
    df['Amount'] = -df['Amount'] #removing all negative signs from values in the dataframe 
    df = df.groupby(['Account'],  as_index=False, sort=False).sum()
    df = pd.concat([pd.DataFrame([['Total', df['Amount'].sum()]], columns=df.columns), df], ignore_index=True) 
    df = pd.concat([pd.DataFrame([['Project Code', f"{project_name}"]], columns=df.columns), df], ignore_index=True) 
    return df

#takes a sheet and returns a df, removes unnecessary columns 
def create_df_list(file, sheet):
    df = pd.read_excel(file, sheet_name = sheet)
    df = df[df['Type'] == 'Bill']

    df = df.drop('Num', axis=1) 
    df = df.drop('Name', axis=1)
    df = df.drop('Type', axis=1)
    df = df.drop('Date', axis=1)
    df = df.drop('Memo', axis=1)

    #create list of dataframes for each project code in file
    df_list = []
    new_df = [] # this is actually a list but will become a dataframe later
    project = df['Project'].iloc[0] # holding project name from first row to keep track for comparisons 


    for i, row in df.iterrows():
        current_project = df['Project'][i]

        if project != current_project:
            append_df = reformat(pd.DataFrame(new_df))
            df_list.append(append_df)
            project = current_project 
            new_df.clear()
        
        new_df.append(row)

    df_list.append(reformat(pd.DataFrame(new_df))) #appending final DataFrame 
    return df_list 

#takes df_list and creates a summary of all of the projects in df_list, returns df
def create_summary(df_list):
    column_values = ['Project Code']
    
    #creating set of unique account types within the worksheet
    for df in df_list:
        for i in range(len(df)):
            account_type = df['Account'].iloc[i] 
            if account_type not in column_values:
                column_values.append(account_type)
    
    column_values.remove('Total') #removing leftover 'Total' value from columns
    
    lst = []
    total_column = [0] #will keep track of the total of each project as they are added to the summary
    df_summary = pd.DataFrame(lst, columns=column_values)
    
    #organize spending from each project code (row) by account type (column)
    for df in df_list:
        lst_append = []
        for t in column_values:
            if t in df['Account'].unique():
                index_value = df.loc[df['Account'] == t].index[0] # index of the corresponding row in df 
                value = df['Amount'].iloc[index_value]
                if isinstance(value, numbers.Number):
                    value = round(value, 2)
                lst_append.append(value)
            else:
                lst_append.append(0) # add a comment about logic here
        df_summary.loc[len(df_summary)] = lst_append
        total_column.append(round(sum(filter(lambda i: isinstance(i, float), lst_append)), 2))
    
    #create and append final row that adds the total spending by account type
    total_row = ['Account Total']
    for t in column_values:
        if t != 'Project Code':
            total_row.append(round(df_summary[t].sum(), 2))

    #insert total row to the top of the summary dataframe
    df_summary.loc[-1] = total_row
    df_summary = df_summary.sort_index().reset_index(drop=True)

    #insert total column into summary dataframe
    df_summary.insert(1, "Project Total", total_column, True)
    total_all_projects = sum(total_column)
    df_summary.at[0, 'Project Total'] = total_all_projects

    return df_summary

#adding rows from dataframe to selected workbook
def append_sheet(wb, sheet_name, df):

    wb.create_sheet(title=sheet_name)

    for r in dataframe_to_rows(df, index=False, header=True):
        wb[sheet_name].append(r)

#Change column width and font of worksheet (Later will add more reformatting changes)
def reformat_workbook(workbook):

    #change column width 
    for sheet in workbook.worksheets:
        dims = {}
        for row in sheet.rows:
            for cell in row:
                #cell style changes
                if cell.row == 1:
                    cell.font = Font(name="Arial",
                                     bold=True,
                                    sz=8)
                else:
                    cell.font = Font(name="Arial",
                                     sz=8)
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                try:    
                    cell.value = int(format (int(cell.value), ',d')) # str(('{:,}'.format(float(cell.value)))) #add commas to cell values 
                except ValueError:
                    continue

        for col, value in dims.items():
            sheet.column_dimensions[col].width = value

#takes an input excel file saves an output file to the directory with the Pro ject cost breakdowns sheet by sheet
def output_file(file):
    wb = load_workbook(file)

    for sheet in wb.worksheets:
        sheet_name = sheet.title
        df_list = create_df_list(file, sheet_name)
        df_summary = create_summary(df_list)
        append_sheet(wb, 'Summary', df_summary) #first add the summary to the workbook

        for df in df_list:
            new_sheet = df['Amount'].iloc[0] #returns the saved project code value from the df 
            append_sheet(wb, new_sheet, df)

        del wb[sheet_name] 
        reformat_workbook(wb)
        wb.save('output_' + file)

files = [f.name for f in pathlib.Path().glob("*.xlsx")]

#incrementing through all files present in folder 
for file in files:
    output_file(file)
